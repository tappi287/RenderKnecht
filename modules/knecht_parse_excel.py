"""
knecht_load_excel for py_knecht. Load excel files

Copyright (C) 2017 Stefan Tapper, All rights reserved.

    This file is part of RenderKnecht Strink Kerker.

    RenderKnecht Strink Kerker is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    RenderKnecht Strink Kerker is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with RenderKnecht Strink Kerker.  If not, see <http://www.gnu.org/licenses/>.

    <renderknecht_varianten> -0-
    |
    |
    |----<origin @sourcedoc>
    |
    |
    |----<model @name @value @modelyear @market @sourcedoc>
    |    |
    |	 |
    |    | # -Farbkombinationen Benutzer Presets oder FakomLutscher Ausgabe-
    |    |----<preset @name @order @type="fakom_setup" @id=int[0:999]>
    |    |    |    |
    |    |    |----<variant @name @order @value />
    |    |    |    |
    |    |----</preset>
    |	 |    |
    |	 |    |
    |	 | # -Serienumfang-
    |	 |----<preset @name @order @type="trim_setup" @id=int[0:999]>
    |	 |    |    |
    |	 |    |----<variant @name @order @value />
    |	 |    |    |
    |	 |----</preset>
    |	 |    |
    |	 |    |
    |	 | # -Packages-
    |	 |----<preset @name @order @type="package" @id=int[0:999]>
    |	 |    |    |
    |	 |    |----<variant @name @order @value />
    |	 |    |	   |
    |	 |----</preset>
    |	 |    |
    |	 |    |
    |	 | # -mögliche Ausstattungsvarianten-
    |	 |----<preset @name @order @type="options" @id=int[0:999]>
    |	 |    |	   |
    |	 |    |----<variant @name @order @value />
    |	 |    |	   |
    |	 |----</preset>
    |    |
    |    |
    |----</model>
    |
    |
    </renderknecht_varianten>
"""

import os
import re
import xml.etree.ElementTree as ET
from PyQt5 import QtCore
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from modules.app_strings import Msg
from modules.app_globals import HELPER_DIR
from modules.knecht_log import init_logging

# Initialize logging for this module
LOGGER = init_logging(__name__)


class LoadVplus(QtCore.QObject):
    """
        Reads V Plus Browser Excel export files and converts them into RenderKnecht readable XML format.
    """
    # Report current status
    status_msg = QtCore.pyqtSignal(str)
    error_msg = QtCore.pyqtSignal(str)

    # Load workbook in read_only mode
    wb_read_only = True

    # Tells class to abort if choosen by user in vplus window
    abort = False
    shorten_names = False
    shorten_pkg_name = False

    # Create trimline presets
    read_trim = True

    # Create presets for optional PR-Codes
    read_options = True

    # PR Familys that will always trigger package read
    read_pkg_options_pr_family_filter = {'SIB', 'VOS', 'KMS'}

    # Create presets for all packages
    read_packages = False

    # Package PR-Family filter
    package_pr_fam_filter = False

    # Create presets for optional packages, no GUI trigger
    read_pkg_options = True

    # PR-Family Filter
    pr_fam_filter = set()

    # Model filter
    model_filter = []

    # PR-Family Storage for lookup in packages
    # eg. pr-option: pr-family
    pr_fam_storage = {'1XW': 'LRA'}

    # Last file name in use
    # Vplus window will store this here, if recently loaded file - apply same filters
    last_file = ''

    def __init__(self, filepath=None):
        # inherit from QObject so we can handle signals
        super(LoadVplus, self).__init__()

        self.filepath = filepath
        self.filename = os.path.basename(self.filepath)

        # Worksheet column definitions, values will be read from title_row + 1
        # attrib dict defines columns to read
        self.ws = dict()

        # Worksheet Coordinates Modelle
        self.ws['model'] = {
            'name': 'Modelle', 'potential_names': ['Modelle', 'Models'], 'title_row': 1,
            # Key to get for descriptive text
            'model_dom_tag': 'model',  # Column names to read attributes from
            'attrib': {  # Modelltext
                'name': 'Modelltext',  # Modell
                'value': 'Modell',  # Modelljahr
                'modelyear': 'Modelljahr',  # Markt
                'market': 'Markt'}}

        # Worksheet Coordinates "PR-Nummern" and "Pakete"
        self.ws['PR'] = {
            'name_pr': 'PR-Nummern', 'potential_names_pr': ['PR-Nummern', 'Interior Scope'],
            'ext_sheet': 'Exterior Scope', 'pr_family_start_row': 6, 'name_pkg': 'Pakete',
            'potential_names_pkg': ['Pakete', 'Packages (purged)'], 'title_row': 3, 'start_model_row': 3,
            'end_model_row': 4, 'start_model_col': 6, 'attrib': {
                # PR-Nummer Name
                'name_col': 3,  # PR-Nummer description
                'desc_col': 4, 'name_start_row': 7, 'pr_family_column': 1}}

        self.xml_dom_tags = dict(root='renderknecht_varianten')

        # Worker variables
        self.pr_fam_filter = set()
        self.ext_copied = False
        self.model = None
        self._book = None

    def read_workbook(self, read_only_mode: bool = False):
        # Load workbook
        self.status_msg.emit(Msg.EXC_FILE_LOAD.format(self.filename))

        with open(self.filepath, 'rb') as f:
            self._book = load_workbook(filename=f, read_only=read_only_mode)

        LOGGER.debug('Workbook loaded: %s', self.filepath)
        self.status_msg.emit(Msg.EXC_FILE_LOADED)

        # Copy "Exterior Scope" worksheet contents to "Interior Scope"
        if not self.ext_copied:
            for sheet_name in self._book.sheetnames:
                if sheet_name == self.ws['PR']['ext_sheet']:
                    self.copy_ext_to_int(wb=self._book, ext_ws=self._book[sheet_name])
                    self.ext_copied = True

        return self._book

    def copy_ext_to_int(self, wb, ext_ws, _int_ws=None):
        """ Copy Exterior Scope rows to Interior Scope """
        temp_rows, temp_cells = [], []

        for sheet_name in wb.sheetnames:
            if sheet_name in self.ws['PR']['potential_names_pr']:
                _int_ws = wb[sheet_name]
                break

        if not _int_ws:
            LOGGER.error('Tried to copy values from "Exterior Scope" to "Interior Scope" but'
                         'could not find "Interior Scope" worksheet.')
            return

        # Read from Exterior Scope worksheet
        for row in ext_ws.iter_rows(min_row=self.ws['PR']['pr_family_start_row']):
            temp_cells = []
            for cell in row:
                temp_cells.append(cell.value)

            # Store exterior cells
            temp_rows.append(temp_cells)

        LOGGER.debug('Adding %s exterior rows to %s interior rows.', ext_ws.max_row, _int_ws.max_row)
        # Append to Interior Scope worksheet
        for row in temp_rows:
            _int_ws.append(row)

        # Create copy of modified workbook
        if self.filepath:
            try:
                name, ext = os.path.splitext(self.filepath)
                new_filename = name + '_modified' + ext
                del ext_ws

                with open(new_filename, 'wb') as f:
                    wb.save(f)
            except Exception as e:
                LOGGER.error('Error saving backup of modified, filtered workbook.\n%s', e)

    def read_worksheet(self, workbook, type_key, sheet_key: str = '', potential_names_key: str = '', worksheet=None):
        """
            Iterate thru existing worksheets and find sheets with name
            matching potential sheet names.
            Update coordinates dictonary sheetname with matching name.
        """
        for sheet_name in workbook.sheetnames:
            if sheet_name in self.ws[type_key][potential_names_key]:
                worksheet = workbook[sheet_name]
                self.ws[type_key][sheet_key] = sheet_name
                break

        if worksheet:
            return worksheet
        else:
            # Report desired worksheet not found
            self.error_msg.emit(Msg.EXC_FILE_WORKSHEET_ERROR.format(self.ws[type_key][sheet_key]))
            return False

    def read_and_return_models(self, wb):
        mod_sheet = self.read_worksheet(wb, 'model', 'name', 'potential_names')

        if not mod_sheet:
            LOGGER.error('Model Worksheet: %s not found.', self.ws['model']['name'])
            return False

        return self.read_model_sheet(mod_sheet, True)

    def read_pr_familys(self, wb):
        """ Read PR familys """
        if not wb:
            self._book = self.read_workbook(LoadVplus.wb_read_only)
        else:
            self._book = wb

        pr_sheet = self.read_worksheet(self._book, 'PR', 'name_pr', 'potential_names_pr')

        if not pr_sheet:
            LOGGER.error('PR Worksheet: %s not found.', self.ws['PR']['name_pr'])
            return False

        if self.ext_copied:
            # Report filtered workbook prepared for reading
            self.status_msg.emit(
                Msg.EXC_FILE_FILTERED_COPIED.format(self.ws['PR']['ext_sheet'], self.ws['PR']['name_pr']))

        # Read PR_Familys from worksheet
        current_row = self.ws['PR']['attrib']['name_start_row'] - 1
        max_row = pr_sheet.max_row
        pr_fam_col = self.ws['PR']['attrib']['pr_family_column']
        temp_pr_fam_set = set()

        while 1:
            current_row += 1
            pr_fam = pr_sheet.cell(row=current_row, column=pr_fam_col).value

            if pr_fam:
                # Read description
                pr_fam_desc = pr_sheet.cell(row=current_row, column=pr_fam_col + 1).value
                # Update PR-Family list
                temp_pr_fam_set.add((pr_fam, pr_fam_desc))

            if current_row >= max_row:
                break

        return temp_pr_fam_set

    def create_document(self, wb):
        self._book = wb
        # Read options given by the dialog
        self.pr_fam_filter = LoadVplus.pr_fam_filter
        self.shorten_names = LoadVplus.shorten_names
        self.read_trim = LoadVplus.read_trim
        self.read_options = LoadVplus.read_options
        self.read_packages = LoadVplus.read_packages
        self.shorten_pkg_name = LoadVplus.shorten_pkg_name

        LOGGER.debug('V Plus read only option: %s', LoadVplus.wb_read_only)

        # Expected sheet names
        potential_sheet_names = self.ws['model']['potential_names'] + self.ws['PR']['potential_names_pr'] + \
                                self.ws['PR']['potential_names_pkg']

        mod_sheet = self.read_worksheet(self._book, 'model', 'name', 'potential_names')
        prn_sheet = self.read_worksheet(self._book, 'PR', 'name_pr', 'potential_names_pr')
        pkg_sheet = self.read_worksheet(self._book, 'PR', 'name_pkg', 'potential_names_pkg')

        # Create sheet name list from just read work sheets
        wb_sheet_names = []
        if mod_sheet:
            wb_sheet_names.append(self.ws['model']['name'])
        if prn_sheet:
            wb_sheet_names.append(self.ws['PR']['name_pr'])
        if pkg_sheet:
            wb_sheet_names.append(self.ws['PR']['name_pkg'])

        if len(wb_sheet_names) == 3:
            LOGGER.debug('Worksheets found: %s', wb_sheet_names)
            self.status_msg.emit(Msg.EXC_FILE_WORKSHEET_FOUND.format(wb_sheet_names))
        else:
            # Abort - not all necessary worksheets have been found
            LOGGER.error('Expected Excel worksheets: %s, detected worksheets: %s',
                         potential_sheet_names,
                         wb_sheet_names)

            self.error_msg.emit(Msg.EXC_FILE_WORKSHEETS_ERROR.format(potential_sheet_names, wb_sheet_names))

            self._book.close()
            del self._book
            return

        # Create xml root
        self.create_xml_root()
        # Create model elements
        self.read_model_sheet(mod_sheet)
        # Create trim_setup elements
        if not self.read_pr_sheet(prn_sheet, pkg_sheet):
            return

        # Return XML Tree file path
        return self.create_xml_tree()

    def read_model_sheet(self, mod_sheet, only_return_list=False):
        """ Reads model definitions from worksheet and creates xml elements accordingly """
        # Modelsheet
        # Iterate Rows with model data
        title_row = self.ws['model']['title_row']
        current_row = title_row
        model_attributes = {}
        model_list = []

        # Detect Model columns, col_index starts at 1
        for col in range(1, 13):
            val = mod_sheet.cell(row=title_row, column=col).value

            # Column index as letter
            column_letter = get_column_letter(col)

            # Find column label and set model attribute
            # to be read from that coord
            if val == self.ws['model']['attrib']['name']:
                self.ws['model']['attrib']['name'] = column_letter
                LOGGER.debug('Found %s in column %s', val, column_letter)
            elif val == self.ws['model']['attrib']['value']:
                self.ws['model']['attrib']['value'] = column_letter
                LOGGER.debug('Found %s in column %s', val, column_letter)
            elif val == self.ws['model']['attrib']['modelyear']:
                self.ws['model']['attrib']['modelyear'] = column_letter
                LOGGER.debug('Found %s in column %s', val, column_letter)
            elif val == self.ws['model']['attrib']['market']:
                self.ws['model']['attrib']['market'] = column_letter
                LOGGER.debug('Found %s in column %s', val, column_letter)

        while 1:
            current_row += 1
            row = str(current_row)
            # Break if empty row reached
            if mod_sheet[self.ws['model']['attrib']['value'] + row].value is None:
                break

            # Read Attributes
            for k, v in self.ws['model']['attrib'].items():
                model_attributes[k] = mod_sheet[v + row].value

            if only_return_list:
                # Append to model_list for model filter list
                model_list.append([model_attributes['value'], model_attributes['name']])
            else:
                # Create Xml element
                self.model = self.create_xml_sub(self.root, 'model', model_attributes)

        if only_return_list:
            return model_list

    def read_pr_sheet(self, prn_sheet, pkg_sheet):
        """ Iterate thru PR and Package worksheets and create presets """

        def shorten_model_name(model_name, num_words: int = 6, shorten: bool = False):
            # Remove horse power description
            model_name = re.sub('(\d?\d\d)[()](...........)\s', '', model_name)
            # Replace to one word
            model_name = re.sub('(S\sline)', 'S-line', model_name)
            model_name = re.sub('(S\stronic)', 'S-tronic', model_name)
            model_name = re.sub('(RS\s)', 'RS', model_name)

            # Split and make sure end index is not smaller than number of words
            # (Do not limit num of words if no shorten set)
            model_name = model_name.split(' ')
            if len(model_name) < num_words or not shorten:
                num_words = len(model_name)

            # If shorten is set, limit to 5 chars/word
            short_name = ''
            for m in model_name[0:num_words]:
                if shorten:
                    short_name += m[0:5] + ' '
                else:
                    short_name += m + ' '

            # Readabilty
            short_name = re.sub('(quatt\s)', 'quattro ', short_name, flags=re.I)
            short_name = re.sub('(Limou\s)', 'Limo ', short_name)
            short_name = re.sub('(allro\s)', 'allroad ', short_name)
            short_name = re.sub('(desig\s)', 'design ', short_name)
            short_name = re.sub('(RSD)', 'RS D', short_name)
            short_name = re.sub('(Navig\s)', 'Navi ', short_name, flags=re.I)
            short_name = re.sub('(Premi\s)', 'Prem ', short_name, flags=re.I)
            short_name = re.sub('(packa\s)', 'Pkg ', short_name, flags=re.I)
            short_name = re.sub('(Techn\s)', 'Tech ', short_name, flags=re.I)
            short_name = re.sub('(Advan\s)', 'Adv ', short_name, flags=re.I)

            return short_name

        # PR Sheet Coordinates
        title_row = self.ws['PR']['title_row']
        min_row = self.ws['PR']['start_model_row']
        max_row = self.ws['PR']['end_model_row'] + 1  # for range iterator +1
        min_col = self.ws['PR']['start_model_col']
        max_col = prn_sheet.max_column
        pr_col = self.ws['PR']['attrib']['name_col']  # PR-Code name column
        pr_start_row = self.ws['PR']['attrib']['name_start_row']  # PR-Code name start row
        pr_family_col = self.ws['PR']['attrib']['pr_family_column']
        desc_col = self.ws['PR']['attrib']['desc_col']  # PR-Description column
        model_dom_tag = self.ws['model']['model_dom_tag']
        current_col = min_col - 1
        model_count = 0
        preset_count = 0

        # Read PR codes
        while 1:
            # Iterate model columns
            current_col += 1
            model_code = ''
            trim_attr = {}
            opti_attr = {}
            opti_pkg_set = {}
            read_model = True

            # Break if empty model column reached
            if prn_sheet.cell(row=min_row, column=current_col).value is None:
                break

            # Create model code from the two model code row's
            for r in range(min_row, max_row):
                model_code += prn_sheet.cell(row=r, column=current_col).value

            if model_code not in self.model_filter: read_model = False

            # LOGGER.debug('Filter: %s Code: %s', self.model_filter, model_code)

            if read_model:
                model_count += 1
                LOGGER.debug('Found model #%s: %s in worksheet %s in column %s', model_count, model_code,
                             self.ws['PR']['name_pr'], current_col)

                # Create trim_setup preset attributes
                # Get model element so we can read attributes from it
                current_model_element = self.root.find(model_dom_tag + '[@value="' + model_code + '"]')

                # Model in PR-Codes but not in Modelsheet
                try:
                    current_market = current_model_element.get('market')
                    model_name = current_model_element.get('name')
                    model_year = current_model_element.get('modelyear')
                    # Preset name: model name, market, model year
                    trim_attr['name'] = shorten_model_name(model_name, 8,
                        self.shorten_names) + current_market + ' ' + model_year

                    # Extract first n string elements eg. Brand Model for options preset name
                    opti_attr['name'] = shorten_model_name(model_name, 3,
                        True) + model_code + ' Options ' + current_market
                except Exception as e:
                    # Model in PR-Codes but not in Modelsheet
                    LOGGER.error('Model %s in PR-Codes but not found in worksheet %s.\n%s',
                                 model_code,
                                 self.ws['model']['name'],
                                 e)
                    break

            trim_attr['value'] = model_code
            opti_attr['value'] = model_code
            trim_attr['type'] = 'trim_setup'
            opti_attr['type'] = 'options'

            # Set trim_setup and options preset's id and order
            if self.read_trim and read_model:
                preset_count += 1
                trim_attr['order'], trim_attr['id'] = str(preset_count), str(preset_count)
                # Create xml tag Trimline Preset
                preset = self.create_xml_sub(current_model_element, 'preset', trim_attr)
                # Create xml tag Modelcode variant
                variant_attr = dict(name=model_code, value='on', order='0')
                variant = self.create_xml_sub(preset, 'variant', variant_attr)

            if self.read_options and read_model:
                preset_count += 1
                opti_attr['order'], opti_attr['id'] = str(preset_count), str(preset_count)
                opt_preset = self.create_xml_sub(current_model_element, 'preset', opti_attr)

            # Iterate PR-Codes
            variant_count = 0
            options_count = 0

            current_row = pr_start_row - 1

            while 1:
                current_row += 1

                # Break if emtpy row reached(new PR-Family begins every third row / 2 empty rows between)
                if prn_sheet.cell(row=current_row, column=pr_col).value is None:
                    current_row += 2
                    if prn_sheet.cell(row=current_row, column=pr_col).value is None:
                        break

                # Find PR-Family
                pr_family = prn_sheet.cell(row=current_row - 1, column=pr_family_col).value
                if pr_family is not None:
                    current_pr_family = pr_family

                # Update PR-Family Storage
                if current_pr_family:
                    pr_opt = prn_sheet.cell(row=current_row, column=pr_col).value
                    if pr_opt:
                        self.pr_fam_storage[pr_opt] = current_pr_family

                # PR Family Filer
                if current_pr_family in self.pr_fam_filter and read_model:
                    # Store trimline codes and options
                    variant_attr = {}
                    variant_attr['type'] = current_pr_family
                    options_attr = {}
                    options_attr['type'] = current_pr_family

                    # Read PR name
                    val = prn_sheet.cell(row=current_row, column=current_col).value

                    if val in ['L', 'I'] and self.read_trim:
                        variant_count += 1
                        variant_attr['name'] = prn_sheet.cell(row=current_row, column=pr_col).value
                        variant_attr['description'] = prn_sheet.cell(row=current_row, column=desc_col).value
                        variant_attr['value'] = 'on'
                        variant_attr['order'] = str(variant_count)

                        # Create xml tag
                        variant = self.create_xml_sub(preset, 'variant', variant_attr)
                    elif val == 'E' and self.read_options:
                        options_count += 1
                        options_attr['name'] = prn_sheet.cell(row=current_row, column=pr_col).value
                        options_attr['description'] = prn_sheet.cell(row=current_row, column=desc_col).value
                        options_attr['value'] = 'on'
                        options_attr['order'] = str(options_count)

                        # Create xml tag
                        option = self.create_xml_sub(opt_preset, 'variant', options_attr)
                    elif val == 'P' and self.read_pkg_options:
                        # Prepare optional packages that will always be read eg. Seats
                        if current_pr_family in self.read_pkg_options_pr_family_filter:
                            name = prn_sheet.cell(row=current_row, column=pr_col).value

                            if current_pr_family in opti_pkg_set.keys():
                                opti_pkg_set[current_pr_family].add(name)
                            else:
                                opti_pkg_set[current_pr_family] = {name}

            # Iterate Packages
            pkg_count = 0

            if (self.read_packages or self.read_pkg_options) and read_model:
                current_row = pr_start_row - 1

                while 2:

                    def create_package_preset(preset_count):
                        options_attr = {}
                        # Values
                        options_attr['value'] = pkg_pr
                        options_attr['type'] = 'package'

                        # Description
                        desc = pkg_sheet.cell(row=current_row - 1, column=pr_family_col + 1).value
                        desc = shorten_model_name(desc, 4, self.shorten_pkg_name)

                        # Name
                        options_attr['name'] = pkg_pr + ' ' + desc + model_code + ' ' + current_market

                        # Order, Id
                        options_attr['order'], options_attr['id'] = str(preset_count), str(preset_count)

                        # Create Xml tag
                        pkg_preset = self.create_xml_sub(current_model_element, 'preset', options_attr)
                        return pkg_preset

                    current_row += 1

                    # Break if empty row reached(new Package begins every third row / 2 empty rows between)
                    if pkg_sheet.cell(row=current_row, column=pr_col).value is None:
                        current_row += 2
                        if pkg_sheet.cell(row=current_row, column=pr_col).value is None:
                            break

                    # Find Package
                    pkg_pr = pkg_sheet.cell(row=current_row - 1, column=pr_family_col).value
                    if pkg_pr is not None:
                        # Set current package
                        current_pkg = pkg_pr
                        # Is it an option?
                        option = pkg_sheet.cell(row=current_row - 1, column=current_col).value
                        skip_package = True

                        if option != '-' and option != 'F' and option is not None:
                            skip_package = False
                            pkg_count += 1
                            preset_count += 1
                            pkg_variant_count = 0
                            pkg_preset = create_package_preset(preset_count)

                    if not skip_package:
                        # Store package codes and options
                        variant_attr = {}

                        # Read PR name
                        val = pkg_sheet.cell(row=current_row, column=current_col).value
                        pr_opt = pkg_sheet.cell(row=current_row, column=pr_col).value

                        # Lookup PR-Family if in storage
                        if pr_opt in self.pr_fam_storage.keys():
                            variant_attr['type'] = self.pr_fam_storage[pr_opt]

                        if val != '-' and val != 'F' and val is not None:
                            pkg_variant_count += 1
                            variant_attr['name'] = pr_opt
                            variant_attr['description'] = pkg_sheet.cell(row=current_row, column=desc_col).value
                            variant_attr['value'] = 'on'
                            variant_attr['order'] = str(pkg_variant_count)

                            # Create xml tag
                            variant = self.create_xml_sub(pkg_preset, 'variant', variant_attr)

                # Iterate packages and remove non-optional ones
                non_essential_msg = list()
                non_fam_msg = list()

                for element in current_model_element.findall('*'):
                    if element.get('type') == 'package':
                        if self.read_pkg_options and not self.read_packages:
                            # Remove non-essential packages and keep eg. Seat-Packages
                            keep_package = False

                            for sub_elem in element.iterfind('.//'):
                                # Iterate PR Familys
                                for items in opti_pkg_set.items():
                                    # Unpack pr_family, options-set eg: VOS, {'Q4H', 'Q1D'}
                                    pr_family, option_set = items

                                    if sub_elem.get('name') in option_set:
                                        keep_package = True
                                        sub_elem.attrib['type'] = pr_family

                            if not keep_package:
                                non_essential_msg.append(element.get('value'))
                        elif self.package_pr_fam_filter and self.read_packages:
                            # Apply optional PR-Family filter to packages
                            keep_package = False

                            # Keep package if any PR-Option matches PR-Family filter
                            for sub_elem in element.iterfind('.//'):
                                if sub_elem.get('type') in self.pr_fam_filter:
                                    keep_package = True

                            if not keep_package:
                                non_fam_msg.append(element.get('value'))

                        if not keep_package:
                            current_model_element.remove(element)

                # Report purged Packages
                if non_essential_msg:
                    LOGGER.debug('Purged non-essential packages %s', non_essential_msg)
                if non_fam_msg:
                    LOGGER.debug('Purged packages, not matching PR-Family filter: %s', non_fam_msg)
                del non_fam_msg, non_essential_msg

            if read_model:
                self.status_msg.emit(Msg.EXC_FILE_MODEL_READ.format(variant_count, pkg_count, model_count, model_code))
                LOGGER.info('Read %s trim variants and %s packages for model #%s: %s.', variant_count, pkg_count,
                    model_count, model_code)
        # All columns parsed
        return True

    def create_xml_tree(self):
        self.pretty_print_xml(self.root)
        self.xmlTree = ET.ElementTree(self.root)

        xmlFilename = os.path.splitext(self.filename)[0] + '.xml'
        xmlPath = HELPER_DIR / xmlFilename

        with open(xmlPath, 'wb') as f:
            self.xmlTree.write(f, encoding='UTF-8', xml_declaration=True)

        return xmlPath

    def create_xml_root(self):
        # XML Root
        self.root = ET.Element(self.xml_dom_tags['root'])

        # Add info about source document
        origin_attr = dict(sourcedoc=self.filename)
        origin = ET.SubElement(self.root, 'origin', origin_attr)

    def create_xml_sub(self, parent, tag, attributes):
        """ Create xml node with given attributes """
        e = ET.SubElement(# Parent Element
            parent, # Tag
            tag, # Attributes to store
            attributes)
        return e

    def pretty_print_xml(self, elem, level=0):
        """ Pretty XML print for better human readabilty """
        NEW_LINE = '\n'
        NEW_LEVEL = '\t'
        i = NEW_LINE + level * NEW_LEVEL

        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = i + NEW_LEVEL
            if not elem.tail or not elem.tail.strip():
                elem.tail = i
            for elem in elem:
                self.pretty_print_xml(elem, level + 1)
            if not elem.tail or not elem.tail.strip():
                elem.tail = i
        else:
            if level and (not elem.tail or not elem.tail.strip()):
                elem.tail = i
